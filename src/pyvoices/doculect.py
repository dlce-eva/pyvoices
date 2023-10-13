"""
Functionality to curate the data collected for one language variety, or doculect.

Processing status of the "raw" data is infered from the existence of derived artefacts.
"""
import pathlib
import collections

import pydub
import pympi
import praatio.textgrid
import openpyxl
from clldutils.misc import lazyproperty


def case_insensitive_glob_pattern(pattern):
    def either(c):
        return '[%s%s]' % (c.lower(), c.upper()) if c.isalpha() else c
    return ''.join(map(either, pattern))


class Doculect:
    def __init__(self, d):
        self.dir = pathlib.Path(d)
        self._eaf_tier_name = None
        self._eaf_tier_names = []

    def _path(self, ext):
        p = '*.' + case_insensitive_glob_pattern(ext)
        try:
            return [p for p in list(self.dir.glob(p)) if not p.name.startswith('~$')][0]
        except:  # pragma: no cover
            raise ValueError(p)

    @property
    def id(self):
        return self._path('wav').stem

    @property
    def eaf_tier_name(self):
        return self._eaf_tier_name

    @property
    def eaf_tier_names(self):
        return self._eaf_tier_names

    @lazyproperty
    def audio(self):
        return pydub.AudioSegment.from_wav(self._path('wav'))

    @lazyproperty
    def praat_textgrid(self):
        try:
            p = self._path('TextGrid')
        except ValueError:
            # Need to convert eaf to TextGrid
            assert self._path('eaf')
            eaf = pympi.Eaf(str(self._path('eaf')))
            tier_names = list(eaf.get_tier_names())
            # If only one tier is found take this otherwise print out all available to rerun it
            if len(tier_names) == 1:
                self._eaf_tier_name = tier_names[0]
            else:
                self._eaf_tier_names = tier_names
            p = self.dir / '{}.TextGrid'.format(self.id)
            eaf.to_textgrid().to_file(str(p))
        return praatio.textgrid.openTextgrid(p, includeEmptyIntervals=False)

    def iter_transcriptions(self, concept_tier='Rfc-Form', concept_column='English Translation'):
        try:
            p = self._path('xlsx')
        except ValueError:
            # No xlsx transcriptions file: seed one with the concept labels from praat:
            p = self.dir / '{}.xlsx'.format(self.id)
            wb = openpyxl.Workbook()
            wb.active.append([self.id, concept_column])
            seen = set()
            for t in self.praat_textgrid.tierDict[concept_tier].entryList:
                if t.label and t.label not in seen:
                    wb.active.append(['', t.label])
                    seen.add(t.label)
            wb.save(p)
        header = None
        wb = openpyxl.load_workbook(p, data_only=True)
        for sname in wb.sheetnames:
            sheet = wb[sname]
            for i, row in enumerate(sheet.rows):
                row = ["" if col.value is None else '{0}'.format(col.value).strip() for col in row]
                if i == 0:
                    header = row
                else:
                    assert header
                    yield collections.OrderedDict(zip(header, row))
            break
